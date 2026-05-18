from io import BytesIO
from collections import OrderedDict
from datetime import datetime
from fastapi import APIRouter, Depends
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import func
from pydantic import BaseModel
from typing import Optional
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from pyexcel_ods3 import save_data as save_ods
from app.database import get_db
from app.models.artesano import Artesano
from app.models.producto import Producto
from app.models.venta import Venta, VentaItem
from app.models.pago_artesano import PagoArtesano
from app.models.ahorro import AhorroArtesano
from app.models.configuracion import Configuracion
from app.services.email import enviar_email, html_pago_artesano

router = APIRouter(prefix="/api/liquidaciones", tags=["liquidaciones"])


class PagoCreate(BaseModel):
    artesano_id: int
    periodo: str
    monto: float
    notas: str = ""


@router.get("/resumen")
def resumen(periodo: Optional[str] = None, db: Session = Depends(get_db)):
    if not periodo:
        hoy = datetime.now()
        periodo = f"{hoy.year}-{hoy.month:02d}"

    year, month = map(int, periodo.split("-"))
    inicio = datetime(year, month, 1, 0, 0, 0)
    if month == 12:
        fin = datetime(year + 1, 1, 1, 0, 0, 0)
    else:
        fin = datetime(year, month + 1, 1, 0, 0, 0)

    ventas = (
        db.query(VentaItem)
        .join(Venta)
        .join(Producto)
        .filter(Venta.fecha >= inicio, Venta.fecha < fin, Venta.estado == "completada")
        .all()
    )

    ventas_por_artesano = {}
    for item in ventas:
        artesano_id = item.producto.artesano_id
        if not artesano_id:
            continue
        if artesano_id not in ventas_por_artesano:
            ventas_por_artesano[artesano_id] = {"vendido": 0, "costo": 0}
        ventas_por_artesano[artesano_id]["vendido"] += item.subtotal
        ventas_por_artesano[artesano_id]["costo"] += item.cantidad * item.producto.costo

    pagos = (
        db.query(
            PagoArtesano.artesano_id,
            func.sum(PagoArtesano.monto).label("total_pagado"),
        )
        .filter(PagoArtesano.periodo == periodo)
        .group_by(PagoArtesano.artesano_id)
        .all()
    )
    pagos_dict = {p.artesano_id: float(p.total_pagado) for p in pagos}

    artesanos = db.query(Artesano).filter(Artesano.activo == True).all()
    resultado = []
    for a in artesanos:
        v = ventas_por_artesano.get(a.id, {"vendido": 0, "costo": 0})
        vendido = round(v["vendido"], 2)
        costo = round(v["costo"], 2)
        deduccion_venta = round(vendido * 0.01, 2)
        deduccion_renta = round(vendido * 0.02, 2)
        deduccion_tienda = round(vendido * 0.02, 2)
        neto = round(vendido - deduccion_venta - deduccion_renta - deduccion_tienda, 2)
        pagado = pagos_dict.get(a.id, 0)
        if vendido > 0 or pagado > 0:
            resultado.append({
                "artesano_id": a.id,
                "artesano": a.nombre,
                "monto_vendido": vendido,
                "deduccion_venta": deduccion_venta,
                "deduccion_renta": deduccion_renta,
                "deduccion_tienda": deduccion_tienda,
                "neto": neto,
                "monto_pagado": round(pagado, 2),
                "pendiente": round(neto - pagado, 2),
            })

    for r in resultado:
        art_id = r["artesano_id"]
        existente = db.query(AhorroArtesano).filter(
            AhorroArtesano.artesano_id == art_id,
            AhorroArtesano.periodo == periodo,
        ).first()
        if not existente:
            ahorro_monto = round(r["monto_vendido"] * 0.05, 2)
            if ahorro_monto > 0:
                db.add(AhorroArtesano(
                    artesano_id=art_id,
                    periodo=periodo,
                    monto_ahorrado=ahorro_monto,
                ))
    db.commit()

    ahorros = db.query(
        AhorroArtesano.artesano_id,
        func.sum(AhorroArtesano.monto_ahorrado).label("total_ahorrado"),
        func.sum(AhorroArtesano.monto_pagado).label("total_pagado_ahorro"),
    ).filter(AhorroArtesano.periodo == periodo).group_by(AhorroArtesano.artesano_id).all()
    ahorros_dict = {a.artesano_id: {"ahorrado": float(a.total_ahorrado), "pagado": float(a.total_pagado_ahorro or 0)} for a in ahorros}

    for r in resultado:
        a = ahorros_dict.get(r["artesano_id"], {"ahorrado": 0, "pagado": 0})
        r["ahorro"] = round(a["ahorrado"], 2)
        r["ahorro_pagado"] = round(a["pagado"], 2)

    return {
        "periodo": periodo,
        "liquidaciones": resultado,
        "total_vendido": round(sum(r["monto_vendido"] for r in resultado), 2),
        "total_neto": round(sum(r["neto"] for r in resultado), 2),
        "total_pagado": round(sum(r["monto_pagado"] for r in resultado), 2),
        "total_pendiente": round(sum(r["pendiente"] for r in resultado), 2),
        "total_ahorro": round(sum(r["ahorro"] for r in resultado), 2),
        "total_ahorro_pagado": round(sum(r["ahorro_pagado"] for r in resultado), 2),
    }


@router.get("/pagos")
def listar_pagos(periodo: Optional[str] = None, db: Session = Depends(get_db)):
    q = db.query(PagoArtesano)
    if periodo:
        q = q.filter(PagoArtesano.periodo == periodo)
    return q.order_by(PagoArtesano.fecha_pago.desc()).all()


class PagoMasivoCreate(BaseModel):
    ids: list[int]
    periodo: str


@router.post("/pagos")
def crear_pago(data: PagoCreate, db: Session = Depends(get_db)):
    p = PagoArtesano(**data.model_dump())
    db.add(p)
    db.commit()
    db.refresh(p)
    return p


@router.post("/pagar-masivo")
def pagar_masivo(data: PagoMasivoCreate, db: Session = Depends(get_db)):
    year, month = map(int, data.periodo.split("-"))
    inicio = datetime(year, month, 1, 0, 0, 0)
    fin = datetime(year + 1, 1, 1, 0, 0, 0) if month == 12 else datetime(year, month + 1, 1, 0, 0, 0)

    negocio = db.query(Configuracion).filter(Configuracion.clave == "nombre_negocio").first()
    nombre_negocio = negocio.valor if negocio else "Mi Negocio"

    ventas = db.query(VentaItem).join(Venta).join(Producto).filter(
        Venta.fecha >= inicio, Venta.fecha < fin, Venta.estado == "completada",
    ).all()

    ventas_por_artesano = {}
    for item in ventas:
        aid = item.producto.artesano_id
        if not aid:
            continue
        if aid not in ventas_por_artesano:
            ventas_por_artesano[aid] = {"vendido": 0}
        ventas_por_artesano[aid]["vendido"] += item.subtotal

    pagos_existentes = db.query(PagoArtesano).filter(
        PagoArtesano.periodo == data.periodo,
        PagoArtesano.artesano_id.in_(data.ids),
    ).all()
    ya_pagados = {p.artesano_id for p in pagos_existentes}

    resultados = []
    for aid in data.ids:
        if aid in ya_pagados:
            resultados.append({"artesano_id": aid, "ok": False, "error": "Ya tiene pago registrado en este período"})
            continue
        v = ventas_por_artesano.get(aid, {"vendido": 0})
        vendido = round(v["vendido"], 2)
        if vendido <= 0:
            resultados.append({"artesano_id": aid, "ok": False, "error": "Sin ventas en el período"})
            continue
        deduccion = round(vendido * 0.05, 2)
        neto = round(vendido - deduccion, 2)

        ahorro_monto = round(vendido * 0.05, 2)
        if ahorro_monto > 0:
            existente = db.query(AhorroArtesano).filter(
                AhorroArtesano.artesano_id == aid,
                AhorroArtesano.periodo == data.periodo,
            ).first()
            if not existente:
                db.add(AhorroArtesano(artesano_id=aid, periodo=data.periodo, monto_ahorrado=ahorro_monto))

        p = PagoArtesano(artesano_id=aid, periodo=data.periodo, monto=neto)
        db.add(p)
        db.flush()

        artesano = db.get(Artesano, aid)
        email_enviado = False
        if artesano and artesano.email:
            html = html_pago_artesano(artesano.nombre, data.periodo, neto, nombre_negocio)
            res = enviar_email(artesano.email, f"Pago registrado - {nombre_negocio}", html)
            email_enviado = res.get("ok", False)

        resultados.append({
            "artesano_id": aid,
            "artesano": artesano.nombre if artesano else f"#{aid}",
            "monto": neto,
            "ok": True,
            "email_enviado": email_enviado,
        })

    db.commit()
    return {"periodo": data.periodo, "pagados": len([r for r in resultados if r["ok"]]), "resultados": resultados}


@router.post("/pagar-todo")
def pagar_todo(periodo: str, db: Session = Depends(get_db)):
    year, month = map(int, periodo.split("-"))
    inicio = datetime(year, month, 1, 0, 0, 0)
    fin = datetime(year + 1, 1, 1, 0, 0, 0) if month == 12 else datetime(year, month + 1, 1, 0, 0, 0)

    ventas = db.query(VentaItem).join(Venta).join(Producto).filter(
        Venta.fecha >= inicio, Venta.fecha < fin, Venta.estado == "completada",
    ).all()

    artesanos_con_ventas = set()
    for item in ventas:
        if item.producto.artesano_id:
            artesanos_con_ventas.add(item.producto.artesano_id)

    ya_pagados = {p.artesano_id for p in db.query(PagoArtesano).filter(
        PagoArtesano.periodo == periodo,
    ).all()}

    pendientes = list(artesanos_con_ventas - ya_pagados)
    if not pendientes:
        return {"ok": True, "pagados": 0, "resultados": []}

    # re-use pagar-masivo
    req = PagoMasivoCreate(ids=pendientes, periodo=periodo)
    return pagar_masivo(req, db)


@router.delete("/pagos/{pago_id}")
def eliminar_pago(pago_id: int, db: Session = Depends(get_db)):
    p = db.get(PagoArtesano, pago_id)
    if not p:
        return {"error": "Pago no encontrado"}
    db.delete(p)
    db.commit()
    return {"ok": True}


@router.post("/ahorros/pagar")
def pagar_ahorro(artesano_id: int, periodo: str, monto: float, db: Session = Depends(get_db)):
    a = db.query(AhorroArtesano).filter(
        AhorroArtesano.artesano_id == artesano_id,
        AhorroArtesano.periodo == periodo,
    ).first()
    if not a:
        return {"error": "No hay ahorros para este artesano en el período"}
    a.monto_pagado = round(a.monto_pagado + monto, 2)
    a.fecha_pago = datetime.now()
    db.commit()
    return {"ok": True, "monto_pagado": a.monto_pagado}


@router.get("/ahorros/historial")
def historial_ahorros(artesano_id: int = None, db: Session = Depends(get_db)):
    q = db.query(AhorroArtesano)
    if artesano_id:
        q = q.filter(AhorroArtesano.artesano_id == artesano_id)
    return q.order_by(AhorroArtesano.periodo.desc()).all()


@router.get("/exportar-excel")
def exportar_excel(periodo: Optional[str] = None, formato: str = "xlsx", db: Session = Depends(get_db)):
    if not periodo:
        hoy = datetime.now()
        periodo = f"{hoy.year}-{hoy.month:02d}"

    year, month = map(int, periodo.split("-"))
    inicio = datetime(year, month, 1, 0, 0, 0)
    if month == 12:
        fin = datetime(year + 1, 1, 1, 0, 0, 0)
    else:
        fin = datetime(year, month + 1, 1, 0, 0, 0)

    ventas = (
        db.query(VentaItem)
        .join(Venta)
        .join(Producto)
        .filter(Venta.fecha >= inicio, Venta.fecha < fin, Venta.estado == "completada")
        .all()
    )

    ventas_por_artesano = {}
    for item in ventas:
        artesano_id = item.producto.artesano_id
        if not artesano_id:
            continue
        if artesano_id not in ventas_por_artesano:
            ventas_por_artesano[artesano_id] = {"vendido": 0}
        ventas_por_artesano[artesano_id]["vendido"] += item.subtotal

    pagos = (
        db.query(
            PagoArtesano.artesano_id,
            func.sum(PagoArtesano.monto).label("total_pagado"),
        )
        .filter(PagoArtesano.periodo == periodo)
        .group_by(PagoArtesano.artesano_id)
        .all()
    )
    pagos_dict = {p.artesano_id: float(p.total_pagado) for p in pagos}

    ahorros = db.query(
        AhorroArtesano.artesano_id,
        func.sum(AhorroArtesano.monto_ahorrado).label("total_ahorrado"),
    ).filter(AhorroArtesano.periodo == periodo).group_by(AhorroArtesano.artesano_id).all()
    ahorros_dict = {a.artesano_id: float(a.total_ahorrado) for a in ahorros}

    artesanos = db.query(Artesano).filter(Artesano.activo == True).all()

    headers = [
        "Código", "Artesano", "Comunidad", "Teléfono",
        "Vendido", "-1% Venta", "-2% Renta", "-2% Tienda",
        "Neto (95%)", "Ya Pagado", "Pendiente", "Ahorro (5%)",
    ]

    rows = []
    total_vendido = total_neto = total_pagado = total_pendiente = total_ahorro = 0

    for a in artesanos:
        v = ventas_por_artesano.get(a.id, {"vendido": 0})
        vendido = round(v["vendido"], 2)
        if vendido == 0 and a.id not in pagos_dict:
            continue
        deduccion_venta = round(vendido * 0.01, 2)
        deduccion_renta = round(vendido * 0.02, 2)
        deduccion_tienda = round(vendido * 0.02, 2)
        neto = round(vendido - deduccion_venta - deduccion_renta - deduccion_tienda, 2)
        pagado = round(pagos_dict.get(a.id, 0), 2)
        pendiente = round(neto - pagado, 2)
        ahorro = round(ahorros_dict.get(a.id, 0), 2)
        comunidad = a.comunidad.nombre if a.comunidad else ""
        rows.append([a.codigo or "", a.nombre, comunidad, a.telefono or "",
                     vendido, deduccion_venta, deduccion_renta, deduccion_tienda,
                     neto, pagado, pendiente, ahorro])
        total_vendido += vendido
        total_neto += neto
        total_pagado += pagado
        total_pendiente += pendiente
        total_ahorro += ahorro

    total_row = ["TOTALES", "", "", "", round(total_vendido, 2), "", "", "",
                 round(total_neto, 2), round(total_pagado, 2), round(total_pendiente, 2), round(total_ahorro, 2)]

    if formato == "ods":
        data = OrderedDict()
        data[f"Liquidaciones {periodo}"] = [headers] + rows + [total_row]
        buf = BytesIO()
        save_ods(buf, data)
        buf.seek(0)
        return StreamingResponse(
            buf,
            media_type="application/vnd.oasis.opendocument.spreadsheet",
            headers={"Content-Disposition": f"attachment; filename=liquidaciones_{periodo}.ods"},
        )

    wb = Workbook()
    ws = wb.active
    ws.title = f"Liquidaciones {periodo}"

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    for i, row in enumerate(rows, 2):
        for j, val in enumerate(row, 1):
            c = ws.cell(row=i, column=j, value=val)
            c.border = thin_border

    fila = len(rows) + 2
    bold_font = Font(bold=True, size=11)
    for j, val in enumerate(total_row, 1):
        c = ws.cell(row=fila, column=j, value=val)
        c.font = bold_font if j in (1, 5, 9, 10, 11, 12) else Font()
        c.border = thin_border

    widths = [12, 25, 20, 14, 14, 12, 12, 12, 14, 14, 14, 14]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=liquidaciones_{periodo}.xlsx"},
    )


@router.get("/historial")
def historial_artesano(artesano_id: int, db: Session = Depends(get_db)):
    pagos = db.query(PagoArtesano).filter(
        PagoArtesano.artesano_id == artesano_id
    ).order_by(PagoArtesano.fecha_pago.desc()).all()

    periodos = sorted(set(p.periodo for p in pagos), reverse=True)
    if not periodos:
        return []

    ventas_por_periodo = {}
    rows = db.query(
        func.strftime("%Y-%m", Venta.fecha).label("periodo"),
        func.sum(VentaItem.cantidad * Producto.costo).label("total"),
    ).select_from(VentaItem).join(Producto).join(Venta).filter(
        Producto.artesano_id == artesano_id,
        Venta.estado == "completada",
    ).group_by(func.strftime("%Y-%m", Venta.fecha)).all()
    for r in rows:
        ventas_por_periodo[r.periodo] = float(r.total or 0)

    pagos_por_periodo = {}
    for p in pagos:
        pagos_por_periodo[p.periodo] = pagos_por_periodo.get(p.periodo, 0) + p.monto

    return [
        {
            "periodo": periodo,
            "monto_vendido": round(ventas_por_periodo.get(periodo, 0), 2),
            "monto_pagado": round(pagos_por_periodo.get(periodo, 0), 2),
            "pendiente": round(ventas_por_periodo.get(periodo, 0) - pagos_por_periodo.get(periodo, 0), 2),
        }
        for periodo in periodos
    ]
