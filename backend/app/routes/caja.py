import os
import json
from io import BytesIO
from collections import OrderedDict
from datetime import datetime, date
from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import func
from pydantic import BaseModel
from typing import Optional, List
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from pyexcel_ods3 import save_data as save_ods
from fpdf import FPDF
from app.database import get_db
from app.models.caja import CierreCaja, MovimientoCaja
from app.models.usuario import Usuario
from app.models.venta import Venta
from app.models.configuracion import Configuracion
from app.services.email import enviar_email, html_cierre

router = APIRouter(prefix="/api/caja", tags=["caja"])


class AperturaCreate(BaseModel):
    usuarios_apertura: List[int]
    monto_inicial_crc: float = 0
    monto_inicial_usd: float = 0


class MovimientoCajaCreate(BaseModel):
    cierre_id: int
    tipo: str
    monto: float
    moneda: str = "CRC"
    descripcion: str = ""


class CierreCreate(BaseModel):
    conteo_crc: float = 0
    conteo_usd: float = 0
    datafono: float = 0
    usuarios_cierre: List[int] = []
    comentarios: Optional[str] = None


def ventas_del_dia(db: Session, apertura: datetime) -> dict:
    crc = db.query(func.coalesce(func.sum(Venta.total), 0)).filter(
        Venta.estado == "completada",
        Venta.fecha >= apertura,
        Venta.moneda == "CRC",
    ).scalar()
    usd = db.query(func.coalesce(func.sum(Venta.total), 0)).filter(
        Venta.estado == "completada",
        Venta.fecha >= apertura,
        Venta.moneda == "USD",
    ).scalar()
    return {"crc": float(crc), "usd": float(usd)}


def obtener_tc(db: Session) -> float:
    tc = db.query(Configuracion).filter(Configuracion.clave == "tipo_cambio_compra").first()
    return float(tc.valor) if tc else 450


def _sum_movs(movimientos, tipo, moneda=None):
    total = 0
    for m in movimientos:
        if m.tipo != tipo:
            continue
        if moneda and m.moneda != moneda:
            continue
        total += m.monto
    return total


def _nombres_usuarios(db: Session, ids: List[int]) -> List[str]:
    if not ids:
        return []
    usuarios = db.query(Usuario).filter(Usuario.id.in_(ids), Usuario.activo == True).all()
    mapa = {u.id: u.nombre for u in usuarios}
    return [mapa.get(uid, f"#{uid}") for uid in ids]


@router.get("/estado")
def estado_caja(db: Session = Depends(get_db)):
    caja = db.query(CierreCaja).filter(CierreCaja.estado == "abierta").first()
    if not caja:
        return {"abierta": False}
    ingresos_crc = _sum_movs(caja.movimientos, "ingreso", "CRC")
    ingresos_usd = _sum_movs(caja.movimientos, "ingreso", "USD")
    egresos_crc = _sum_movs(caja.movimientos, "egreso", "CRC")
    egresos_usd = _sum_movs(caja.movimientos, "egreso", "USD")
    depositos_crc = _sum_movs(caja.movimientos, "deposito", "CRC")
    depositos_usd = _sum_movs(caja.movimientos, "deposito", "USD")
    v = ventas_del_dia(db, caja.fecha_apertura)
    esperado_crc = caja.monto_inicial_crc + v["crc"] + ingresos_crc - egresos_crc - depositos_crc
    esperado_usd = caja.monto_inicial_usd + v["usd"] + ingresos_usd - egresos_usd - depositos_usd
    ids_ap = caja.apertura_usuarios
    return {
        "abierta": True,
        "id": caja.id,
        "usuarios_apertura": ids_ap,
        "usuarios_apertura_nombres": _nombres_usuarios(db, ids_ap),
        "fecha_apertura": caja.fecha_apertura,
        "monto_inicial_crc": caja.monto_inicial_crc,
        "monto_inicial_usd": caja.monto_inicial_usd,
        "ventas_crc": v["crc"],
        "ventas_usd": v["usd"],
        "ingresos_crc": ingresos_crc,
        "ingresos_usd": ingresos_usd,
        "egresos_crc": egresos_crc,
        "egresos_usd": egresos_usd,
        "depositos_crc": depositos_crc,
        "depositos_usd": depositos_usd,
        "esperado_crc": esperado_crc,
        "esperado_usd": esperado_usd,
        "conteo_crc": caja.conteo_crc,
        "conteo_usd": caja.conteo_usd,
        "datafono": caja.datafono,
        "diferencia_crc": caja.diferencia_crc,
        "diferencia_usd": caja.diferencia_usd,
        "comentarios": caja.comentarios,
        "movimientos": caja.movimientos,
    }


@router.post("/abrir")
def abrir_caja(data: AperturaCreate, db: Session = Depends(get_db)):
    abierta = db.query(CierreCaja).filter(CierreCaja.estado == "abierta").first()
    if abierta:
        return {"error": "Ya hay una caja abierta"}
    c = CierreCaja(
        usuario_id=data.usuarios_apertura[0] if data.usuarios_apertura else 1,
        usuarios_apertura=json.dumps(data.usuarios_apertura),
        monto_inicial_crc=data.monto_inicial_crc,
        monto_inicial_usd=data.monto_inicial_usd,
    )
    db.add(c)
    db.commit()
    db.refresh(c)
    return c


@router.post("/cerrar")
def cerrar_caja(data: CierreCreate, db: Session = Depends(get_db)):
    caja = db.query(CierreCaja).filter(CierreCaja.estado == "abierta").first()
    if not caja:
        return {"error": "No hay caja abierta"}
    ingresos_crc = _sum_movs(caja.movimientos, "ingreso", "CRC")
    egresos_crc = _sum_movs(caja.movimientos, "egreso", "CRC")
    depositos_crc = _sum_movs(caja.movimientos, "deposito", "CRC")
    depositos_usd = _sum_movs(caja.movimientos, "deposito", "USD")
    v = ventas_del_dia(db, caja.fecha_apertura)
    esperado_crc = caja.monto_inicial_crc + v["crc"] + ingresos_crc - egresos_crc - depositos_crc
    esperado_usd = caja.monto_inicial_usd + v["usd"] - depositos_usd
    caja.conteo_crc = data.conteo_crc
    caja.conteo_usd = data.conteo_usd
    caja.datafono = data.datafono
    if data.usuarios_cierre:
        caja.usuario_cierre_id = data.usuarios_cierre[0]
        caja.usuarios_cierre = json.dumps(data.usuarios_cierre)
    caja.comentarios = data.comentarios
    caja.diferencia_crc = round(data.conteo_crc - esperado_crc, 2)
    caja.diferencia_usd = round(data.conteo_usd - esperado_usd, 2)
    caja.fecha_cierre = datetime.now()
    caja.estado = "cerrada"
    db.commit()
    return caja


@router.post("/movimiento")
def registrar_movimiento(data: MovimientoCajaCreate, db: Session = Depends(get_db)):
    m = MovimientoCaja(**data.model_dump())
    db.add(m)
    db.commit()
    return m


def _serializar_historial(c, db):
    ids_ap = c.apertura_usuarios
    ids_cr = c.cierre_usuarios if c.fecha_cierre else []
    return {
        "id": c.id,
        "fecha_apertura": c.fecha_apertura,
        "fecha_cierre": c.fecha_cierre,
        "usuarios_apertura": ids_ap,
        "usuarios_apertura_nombres": _nombres_usuarios(db, ids_ap),
        "usuarios_cierre": ids_cr,
        "usuarios_cierre_nombres": _nombres_usuarios(db, ids_cr),
        "monto_inicial_crc": c.monto_inicial_crc,
        "monto_inicial_usd": c.monto_inicial_usd,
        "conteo_crc": c.conteo_crc,
        "conteo_usd": c.conteo_usd,
        "datafono": c.datafono,
        "diferencia_crc": c.diferencia_crc,
        "diferencia_usd": c.diferencia_usd,
        "comentarios": c.comentarios,
    }


@router.get("/historial")
def historial(db: Session = Depends(get_db)):
    cierres = db.query(CierreCaja).order_by(CierreCaja.fecha_apertura.desc()).limit(50).all()
    return [_serializar_historial(c, db) for c in cierres]


def _estado_cierre_dict(c: CierreCaja, db: Session) -> dict:
    ingresos_crc = _sum_movs(c.movimientos, "ingreso", "CRC")
    ingresos_usd = _sum_movs(c.movimientos, "ingreso", "USD")
    egresos_crc = _sum_movs(c.movimientos, "egreso", "CRC")
    egresos_usd = _sum_movs(c.movimientos, "egreso", "USD")
    depositos_crc = _sum_movs(c.movimientos, "deposito", "CRC")
    depositos_usd = _sum_movs(c.movimientos, "deposito", "USD")
    v = ventas_del_dia(db, c.fecha_apertura)
    ids_ap = c.apertura_usuarios
    ids_cr = c.cierre_usuarios if c.fecha_cierre else []
    return {
        "id": c.id,
        "fecha_apertura": c.fecha_apertura,
        "fecha_cierre": c.fecha_cierre,
        "usuarios_apertura": ids_ap,
        "usuarios_apertura_nombres": _nombres_usuarios(db, ids_ap),
        "usuarios_cierre": ids_cr,
        "usuarios_cierre_nombres": _nombres_usuarios(db, ids_cr),
        "monto_inicial_crc": c.monto_inicial_crc,
        "monto_inicial_usd": c.monto_inicial_usd,
        "ventas_crc": v["crc"],
        "ventas_usd": v["usd"],
        "ingresos_crc": ingresos_crc,
        "ingresos_usd": ingresos_usd,
        "egresos_crc": egresos_crc,
        "egresos_usd": egresos_usd,
        "depositos_crc": depositos_crc,
        "depositos_usd": depositos_usd,
        "esperado_crc": c.monto_inicial_crc + v["crc"] + ingresos_crc - egresos_crc - depositos_crc,
        "esperado_usd": c.monto_inicial_usd + v["usd"] + ingresos_usd - egresos_usd - depositos_usd,
        "conteo_crc": c.conteo_crc,
        "conteo_usd": c.conteo_usd,
        "datafono": c.datafono,
        "diferencia_crc": c.diferencia_crc,
        "diferencia_usd": c.diferencia_usd,
        "comentarios": c.comentarios,
        "movimientos": c.movimientos,
    }


@router.get("/{cierre_id}/exportar-excel")
def exportar_excel_cierre(cierre_id: int, formato: str = Query("xlsx", regex="^(xlsx|ods)$"), db: Session = Depends(get_db)):
    c = db.get(CierreCaja, cierre_id)
    if not c:
        return {"error": "Cierre no encontrado"}
    est = _estado_cierre_dict(c, db)

    headers = ["Concepto", "CRC", "USD"]
    def fila(label, crc, usd=""):
        return [label, f"₡{crc:,.2f}" if crc else "", f"${usd:,.2f}" if usd else ""]
    rows = [
        fila("Inicial", est["monto_inicial_crc"], est["monto_inicial_usd"]),
        fila("Ventas", est["ventas_crc"], est["ventas_usd"]),
        fila("Ingresos", est["ingresos_crc"]),
        fila("Egresos", est["egresos_crc"]),
        fila("Depósitos banco", est["depositos_crc"], est["depositos_usd"]),
        fila("Esperado", est["esperado_crc"], est["esperado_usd"]),
        fila("Conteo final", est.get("conteo_crc", 0) or 0, est.get("conteo_usd", 0) or 0),
        fila("Datáfono", est.get("datafono", 0) or 0),
        fila("Diferencia", est.get("diferencia_crc", 0) or 0, est.get("diferencia_usd", 0) or 0),
    ]

    mov_headers = ["Tipo", "Moneda", "Monto", "Descripción"]
    mov_rows = []
    for m in c.movimientos:
        mon = "$" if m.moneda == "USD" else "₡"
        mov_rows.append([m.tipo.capitalize(), m.moneda or "CRC", f"{mon}{m.monto:,.2f}", m.descripcion or "-"])

    if formato == "ods":
        data = OrderedDict()
        data["Resumen"] = [headers] + rows
        if mov_rows:
            data["Movimientos"] = [mov_headers] + mov_rows
        buf = BytesIO()
        save_ods(buf, data)
        buf.seek(0)
        return StreamingResponse(buf, media_type="application/vnd.oasis.opendocument.spreadsheet",
            headers={"Content-Disposition": f"attachment; filename=cierre_{cierre_id}.ods"})

    wb = Workbook()
    ws = wb.active
    ws.title = "Resumen"
    hf = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.fill = hf
        cell.alignment = Alignment(horizontal="center")
    for ri, row in enumerate(rows, 2):
        for ci, val in enumerate(row, 1):
            ws.cell(row=ri, column=ci, value=val)
    if mov_rows:
        ws2 = wb.create_sheet("Movimientos")
        for ci, h in enumerate(mov_headers, 1):
            cell = ws2.cell(row=1, column=ci, value=h)
            cell.font = Font(bold=True, color="FFFFFF", size=11)
            cell.fill = hf
            cell.alignment = Alignment(horizontal="center")
        for ri, row in enumerate(mov_rows, 2):
            for ci, val in enumerate(row, 1):
                ws2.cell(row=ri, column=ci, value=val)
    for i, w in enumerate([20, 20, 20], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=cierre_{cierre_id}.xlsx"})


_FONT_CANDIDATES = [
    "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
    "/usr/share/fonts/TTF/DejaVuSans.ttf",
]
_FONT_DIR = next((os.path.dirname(p) for p in _FONT_CANDIDATES if os.path.exists(p)), "/usr/share/fonts/TTF")
FONT_REGULAR = f"{_FONT_DIR}/DejaVuSans.ttf"
FONT_BOLD = f"{_FONT_DIR}/DejaVuSans-Bold.ttf"


@router.get("/{cierre_id}/pdf")
def exportar_pdf_cierre(cierre_id: int, db: Session = Depends(get_db)):
    c = db.get(CierreCaja, cierre_id)
    if not c:
        return {"error": "Cierre no encontrado"}
    est = _estado_cierre_dict(c, db)
    neg = db.query(Configuracion).filter(Configuracion.clave == "nombre_negocio").first()
    nombre_negocio = neg.valor if neg else "Mi Negocio"
    ap = ", ".join(est["usuarios_apertura_nombres"]) or "-"
    cr = ", ".join(est["usuarios_cierre_nombres"]) or "-"

    pdf = FPDF()
    pdf.add_font("DejaVu", "", FONT_REGULAR, uni=True)
    pdf.add_font("DejaVu", "B", FONT_BOLD, uni=True)
    pdf.add_page()
    pdf.set_font("DejaVu", "B", 14)
    pdf.cell(0, 10, "Cierre de Caja", new_x="LMARGIN", new_y="NEXT")
    pdf.set_font("DejaVu", "", 10)
    pdf.cell(0, 6, nombre_negocio, new_x="LMARGIN", new_y="NEXT")
    pdf.cell(0, 6, f"Apertura: {est['fecha_apertura'].strftime('%d/%m/%Y %H:%M')}", new_x="LMARGIN", new_y="NEXT")
    pdf.cell(0, 6, f"Cierre: {est['fecha_cierre'].strftime('%d/%m/%Y %H:%M') if est['fecha_cierre'] else '-'}", new_x="LMARGIN", new_y="NEXT")
    pdf.cell(0, 6, f"Abrio: {ap}", new_x="LMARGIN", new_y="NEXT")
    pdf.cell(0, 6, f"Cerro: {cr}", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(4)
    cols = [50, 35, 35]
    pdf.set_font("DejaVu", "B", 9)
    pdf.set_fill_color(220, 220, 220)
    for i, h in enumerate(["Concepto", "CRC", "USD"]):
        pdf.cell(cols[i], 7, h, border=1, fill=True, align="C")
    pdf.ln()
    pdf.set_font("DejaVu", "", 8)
    def pdf_fila(label, crc, usd=""):
        pdf.cell(cols[0], 6, label, border=1)
        pdf.cell(cols[1], 6, f"₡{crc:,.2f}" if crc else "", border=1, align="R")
        pdf.cell(cols[2], 6, f"${usd:,.2f}" if usd else "", border=1, align="R")
        pdf.ln()
    pdf_fila("Inicial", est["monto_inicial_crc"], est["monto_inicial_usd"])
    pdf_fila("Ventas", est["ventas_crc"], est["ventas_usd"])
    pdf_fila("Ingresos", est["ingresos_crc"])
    pdf_fila("Egresos", est["egresos_crc"])
    pdf_fila("Depositos banco", est["depositos_crc"], est["depositos_usd"])
    pdf_fila("Esperado", est["esperado_crc"], est["esperado_usd"])
    pdf_fila("Conteo final", est.get("conteo_crc", 0) or 0, est.get("conteo_usd", 0) or 0)
    pdf_fila("Datafono", est.get("datafono", 0) or 0)
    pdf_fila("Diferencia", est.get("diferencia_crc", 0) or 0, est.get("diferencia_usd", 0) or 0)
    if est.get("comentarios"):
        pdf.ln(4)
        pdf.set_font("DejaVu", "", 8)
        pdf.multi_cell(0, 5, f"Comentarios: {est['comentarios']}")
    buf = BytesIO()
    pdf.output(buf)
    buf.seek(0)
    return StreamingResponse(buf, media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=cierre_{cierre_id}.pdf"})


@router.post("/{cierre_id}/enviar-correo")
def enviar_cierre_correo(cierre_id: int, destinatario: str = "", db: Session = Depends(get_db)):
    c = db.get(CierreCaja, cierre_id)
    if not c:
        return {"error": "Cierre no encontrado"}

    email = destinatario.strip()
    if not email:
        cfg = db.query(Configuracion).filter(Configuracion.clave == "correo_cierre").first()
        email = cfg.valor.strip() if cfg and cfg.valor else ""

    if not email:
        return {"error": "No hay correo destino. Configurá Correo del cierre o escribí uno"}

    neg = db.query(Configuracion).filter(Configuracion.clave == "nombre_negocio").first()
    nombre_negocio = neg.valor if neg else "Mi Negocio"
    est = _estado_cierre_dict(c, db)
    html = html_cierre(est, nombre_negocio)
    resultado = enviar_email(email, f"Cierre de caja #{cierre_id} - {nombre_negocio}", html)
    return resultado
