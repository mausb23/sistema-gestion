import json
from collections import OrderedDict
from datetime import datetime, date
from fastapi import APIRouter, Depends
from sqlalchemy.orm import Session
from sqlalchemy import func
from pydantic import BaseModel
from typing import Optional, List
from fastapi.responses import StreamingResponse
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from pyexcel_ods3 import save_data as save_ods
from app.database import get_db
from app.models.venta import Venta, VentaItem
from app.models.producto import Producto
from app.models.inventario import MovimientoInventario
from app.models.configuracion import Configuracion

router = APIRouter(prefix="/api/ventas", tags=["ventas"])


class ItemInput(BaseModel):
    producto_id: int
    cantidad: float
    precio_unitario: float


class PagoInput(BaseModel):
    metodo: str
    monto: float
    moneda: str = "CRC"


class VentaCreate(BaseModel):
    usuario_id: int
    cliente_id: Optional[int] = None
    moneda: str = "CRC"
    metodo_pago: str = "efectivo"
    pagos: List[PagoInput] = []
    items: List[ItemInput]


@router.get("")
def listar(fecha_desde: Optional[str] = None, fecha_hasta: Optional[str] = None, db: Session = Depends(get_db)):
    q = db.query(Venta)
    if fecha_desde:
        q = q.filter(Venta.fecha >= datetime.fromisoformat(fecha_desde))
    if fecha_hasta:
        q = q.filter(Venta.fecha <= datetime.fromisoformat(fecha_hasta))
    return q.order_by(Venta.fecha.desc()).all()


@router.get("/hoy")
def ventas_hoy(db: Session = Depends(get_db)):
    tc = db.query(Configuracion).filter(Configuracion.clave == "tipo_cambio_compra").first()
    tc_compra = float(tc.valor) if tc else 500

    hoy = date.today()
    inicio = datetime(hoy.year, hoy.month, hoy.day, 0, 0, 0)
    fin = datetime(hoy.year, hoy.month, hoy.day, 23, 59, 59)
    q = db.query(Venta).filter(Venta.fecha >= inicio, Venta.fecha <= fin)
    ventas = q.order_by(Venta.fecha.desc()).all()
    total = sum(v.total for v in ventas)

    resumen_pagos = {}
    for v in ventas:
        if v.estado != "completada":
            continue
        if v.pagos_detalle:
            try:
                pagos = json.loads(v.pagos_detalle)
                for pago in pagos:
                    metodo = pago.get("metodo", v.metodo_pago)
                    monto = pago.get("monto", 0)
                    moneda = pago.get("moneda", "CRC")
                    if moneda == "USD":
                        monto = round(monto * tc_compra, 2)
                    resumen_pagos[metodo] = resumen_pagos.get(metodo, 0) + monto
            except (json.JSONDecodeError, TypeError):
                key = v.metodo_pago
                if key == "efectivo" and v.moneda == "USD":
                    key = "efectivo_dolares"
                resumen_pagos[key] = resumen_pagos.get(key, 0) + v.total
        else:
            key = v.metodo_pago
            if key == "efectivo" and v.moneda == "USD":
                key = "efectivo_dolares"
            resumen_pagos[key] = resumen_pagos.get(key, 0) + v.total

    return {"ventas": ventas, "total": total, "cantidad": len(ventas), "resumen_pagos": resumen_pagos}


@router.post("")
def crear(data: VentaCreate, db: Session = Depends(get_db)):
    total = sum(item.cantidad * item.precio_unitario for item in data.items)
    pagos_detalle = None
    metodo_pago = data.metodo_pago
    if data.pagos:
        pagos_detalle = json.dumps([p.model_dump() for p in data.pagos])
        metodos = "+".join(sorted(set(p.metodo for p in data.pagos)))
        monedas = set(p.moneda for p in data.pagos)
        if "CRC" in monedas and "USD" in monedas:
            moneda = "CRC"
        elif "USD" in monedas:
            moneda = "USD"
        metodo_pago = metodos

    venta = Venta(
        usuario_id=data.usuario_id,
        cliente_id=data.cliente_id,
        moneda=data.moneda,
        metodo_pago=metodo_pago,
        total=total,
        pagos_detalle=pagos_detalle,
    )
    db.add(venta)
    db.flush()

    for item in data.items:
        vi = VentaItem(
            venta_id=venta.id,
            producto_id=item.producto_id,
            cantidad=item.cantidad,
            precio_unitario=item.precio_unitario,
            subtotal=item.cantidad * item.precio_unitario,
        )
        db.add(vi)

        producto = db.query(Producto).get(item.producto_id)
        if producto:
            nuevo_stock = max(0, producto.stock - item.cantidad)
            producto.stock = nuevo_stock
            db.add(MovimientoInventario(
                producto_id=item.producto_id,
                tipo="salida",
                cantidad=item.cantidad,
                stock_resultante=nuevo_stock,
                motivo=f"Venta #{venta.id}",
                usuario_id=data.usuario_id,
            ))

    db.commit()
    db.refresh(venta)
    return venta


@router.delete("/{venta_id}")
def eliminar(venta_id: int, db: Session = Depends(get_db)):
    v = db.query(Venta).get(venta_id)
    if not v:
        return {"error": "Venta no encontrada"}
    for item in v.items:
        producto = db.query(Producto).get(item.producto_id)
        if producto:
            producto.stock += item.cantidad
    v.estado = "anulada"
    db.commit()
    return {"ok": True}


@router.get("/exportar-excel")
def exportar_excel(fecha_desde: Optional[str] = None, fecha_hasta: Optional[str] = None, formato: str = "xlsx", db: Session = Depends(get_db)):
    tc = db.query(Configuracion).filter(Configuracion.clave == "tipo_cambio_compra").first()
    tc_compra = float(tc.valor) if tc else 450

    q = db.query(Venta).filter(Venta.estado == "completada").order_by(Venta.fecha)
    if fecha_desde:
        q = q.filter(Venta.fecha >= datetime.fromisoformat(fecha_desde))
    if fecha_hasta:
        q = q.filter(Venta.fecha <= datetime.fromisoformat(fecha_hasta))
    ventas = q.all()

    headers = ["Fecha", "Artículo", "Cantidad", "Costo Unitario", "Nombre Artesano",
               "Número Comunidad", "Monto Total", "Método de Pago", "Dólares recibidos", "Detalles y comentarios"]

    rows = []
    for v in ventas:
        for item in v.items:
            p = item.producto
            comunidad = p.artesano.comunidad if p and p.artesano else None
            comunidad_numero = f"{comunidad.codigo} - {comunidad.nombre}" if comunidad else ""
            metodo = v.metodo_pago
            if metodo == "efectivo_dolares":
                metodo = "Efectivo Dólares"
            elif metodo == "sinpe":
                metodo = "SINPE Móvil"
            dolares = round(item.subtotal / tc_compra, 2) if v.moneda == "USD" else ""
            rows.append([
                v.fecha.strftime("%d/%m/%Y %H:%M"),
                p.nombre if p else "",
                item.cantidad,
                p.costo if p else 0,
                p.artesano.nombre if p and p.artesano else "",
                comunidad_numero,
                round(item.subtotal, 2),
                metodo,
                dolares,
                "",
            ])

    if formato == "ods":
        data = OrderedDict()
        data["Ventas"] = [headers] + rows
        buf = BytesIO()
        save_ods(buf, data)
        buf.seek(0)
        return StreamingResponse(
            buf,
            media_type="application/vnd.oasis.opendocument.spreadsheet",
            headers={"Content-Disposition": "attachment; filename=ventas.ods"},
        )

    wb = Workbook()
    ws = wb.active
    ws.title = "Ventas"

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    for i, row in enumerate(rows, 2):
        for j, val in enumerate(row, 1):
            ws.cell(row=i, column=j, value=val).border = thin_border

    widths = [18, 30, 10, 15, 25, 25, 15, 20, 18, 25]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + i)].width = w

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=ventas.xlsx"},
    )
