import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent / "backend"))

from openpyxl import load_workbook
from app.database import SessionLocal, init_db
from app.models.producto import Producto
from app.models.artesano import Artesano
from app.models.comunidad import Comunidad

EXCEL_PATH = Path(__file__).resolve().parent.parent / "migracion.xlsx"
SHEET = "Inventarios"
BATCH_SIZE = 200


def limpiar_nombre(raw: str) -> str:
    raw = raw.strip()
    parts = raw.split(" (")
    if len(parts) > 1:
        raw = parts[0]
    parts = raw.split(" (")
    raw = parts[0] if len(parts) > 1 else raw
    raw = raw.split(" cta")[0]
    raw = raw.split(" CTA")[0]
    raw = raw.split(" Cuenta")[0]
    raw = raw.split(" cuenta")[0]
    raw = raw.split(" SINPE")[0]
    raw = raw.split(" sinpe")[0]
    raw = raw.split("  ")[0]
    return raw.strip().upper()


def main():
    init_db()
    db = SessionLocal()

    wb = load_workbook(EXCEL_PATH, data_only=True)
    ws = wb[SHEET]

    comunidad_cache = {}
    for c in db.query(Comunidad).all():
        comunidad_cache[c.codigo] = c.id

    artesano_cache = {}
    for a in db.query(Artesano).all():
        artesano_cache[a.nombre.upper()] = a.id

    codigos_existentes = set()
    for p in db.query(Producto.codigo).all():
        codigos_existentes.add(p.codigo)

    insertados = 0
    saltados = 0
    batch = []

    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 1):
        codigo = str(row[0]).strip() if row[0] else ""
        if not codigo or codigo == "None":
            saltados += 1
            continue

        if codigo in codigos_existentes:
            saltados += 1
            continue

        nombre = str(row[3]).strip() if row[3] else ""
        if not nombre or nombre == "None":
            saltados += 1
            continue
        nombre = nombre[:200]

        comunidad_codigo = str(row[1]).strip() if row[1] else ""
        descripcion = str(row[17]).strip() if row[17] and row[17] != "None" else ""
        descripcion = descripcion[:500]

        try:
            costo_raw = float(row[10]) if row[10] else 0
        except (ValueError, TypeError):
            costo_raw = 0
        try:
            precio_raw = float(row[11]) if row[11] else 0
        except (ValueError, TypeError):
            precio_raw = 0
        try:
            stock_raw = float(row[9]) if row[9] is not None else 0
        except (ValueError, TypeError):
            stock_raw = 0
        if stock_raw < 0:
            stock_raw = 0

        artesano_raw = str(row[4]).strip() if row[4] else ""
        artesano_id = None
        if artesano_raw and artesano_raw != "None":
            nombre_limpio = limpiar_nombre(artesano_raw)
            if nombre_limpio in artesano_cache:
                artesano_id = artesano_cache[nombre_limpio]
            else:
                comunidad_id = None
                if comunidad_codigo in comunidad_cache:
                    comunidad_id = comunidad_cache[comunidad_codigo]
                numero_art = len([x for x in artesano_cache.values()]) + 1
                art = Artesano(
                    codigo=f"{comunidad_codigo}-{str(numero_art).zfill(3)}",
                    nombre=nombre_limpio,
                    comunidad_id=comunidad_id,
                )
                db.add(art)
                db.flush()
                artesano_id = art.id
                artesano_cache[nombre_limpio] = art.id

        p = Producto(
            codigo=codigo,
            nombre=nombre,
            descripcion=descripcion,
            artesano_id=artesano_id,
            precio=precio_raw,
            costo=costo_raw,
            stock=stock_raw,
            stock_minimo=0,
        )
        batch.append(p)
        codigos_existentes.add(codigo)

        if len(batch) >= BATCH_SIZE:
            db.add_all(batch)
            db.commit()
            insertados += len(batch)
            batch = []
            print(f"  Progreso: {insertados} insertados, {saltados} saltados...")

    if batch:
        db.add_all(batch)
        db.commit()
        insertados += len(batch)

    db.close()
    wb.close()
    print(f"\n Migración completa: {insertados} insertados, {saltados} saltados")


if __name__ == "__main__":
    main()
