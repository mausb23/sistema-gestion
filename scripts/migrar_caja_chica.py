import sys
from pathlib import Path
from datetime import datetime

sys.path.insert(0, str(Path(__file__).resolve().parent.parent / "backend"))

from openpyxl import load_workbook
from app.database import SessionLocal, init_db
from app.models.venta import Venta, VentaItem
from app.models.producto import Producto
from app.models.usuario import Usuario

EXCEL_PATH = Path(__file__).resolve().parent.parent / "migracion.xlsx"
SHEET = "Caja Chica"

MAPA_METODO = {
    "tarjeta": "tarjeta",
    "efectivo colones": "efectivo",
    "efectivo dólares": "efectivo_dolares",
}


def main():
    init_db()
    db = SessionLocal()

    usuario = db.query(Usuario).first()
    if not usuario:
        usuario = Usuario(nombre="Admin", rol="admin")
        db.add(usuario)
        db.flush()

    codigos_productos = {}
    for p in db.query(Producto).all():
        codigos_productos[p.codigo] = p

    wb = load_workbook(EXCEL_PATH, data_only=True)
    ws = wb[SHEET]

    insertados = 0
    saltados = 0

    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
        fecha_raw = row[0]
        codigo = str(row[1]).strip() if row[1] else ""
        cant_raw = row[3]
        cantidad = 1
        if cant_raw is not None:
            try:
                cantidad = float(str(cant_raw).strip())
            except (ValueError, TypeError):
                cantidad = 1
        monto = 0
        if row[8] is not None:
            try:
                monto = float(str(row[8]).strip())
            except (ValueError, TypeError):
                monto = 0
        metodo_raw = str(row[9]).strip().lower() if row[9] else ""
        dolares = row[10]

        if not codigo or codigo == "None":
            continue

        if codigo not in codigos_productos:
            saltados += 1
            continue

        producto = codigos_productos[codigo]
        metodo = MAPA_METODO.get(metodo_raw, metodo_raw)
        moneda = "USD" if metodo == "efectivo_dolares" else "CRC"

        if isinstance(fecha_raw, datetime):
            fecha = fecha_raw
        elif fecha_raw:
            fecha = datetime.fromisoformat(str(fecha_raw))
        else:
            fecha = datetime.now()

        precio_unitario = monto / cantidad if cantidad > 0 else producto.precio
        subtotal = monto

        venta = Venta(
            usuario_id=usuario.id,
            fecha=fecha,
            total=subtotal,
            moneda=moneda,
            metodo_pago=metodo,
            estado="completada",
        )
        db.add(venta)
        db.flush()

        vi = VentaItem(
            venta_id=venta.id,
            producto_id=producto.id,
            cantidad=cantidad,
            precio_unitario=precio_unitario,
            subtotal=subtotal,
        )
        db.add(vi)

        insertados += 1

        if insertados % 20 == 0:
            db.commit()
            print(f"  Progreso: {insertados} insertados, {saltados} saltados...")

    db.commit()
    db.close()
    wb.close()
    print(f"\n Migración de ventas completa: {insertados} insertadas, {saltados} saltadas")


if __name__ == "__main__":
    main()
